from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_protect

import datetime, pytz
from datetime import timedelta
from django.utils import timezone

from django.db.models import Sum, Avg, Count
from django.db.models.functions import TruncDay 

import xlwt

import openpyxl
from openpyxl.styles.borders import Border,Side
from django.contrib.gis.geoip2 import GeoIP2
from api.models import *
from . check_time_zone import show_user_time
import pytz
from datetime import datetime as dt
from itertools import zip_longest
import time

from json import dumps
from django.db.models import Q
import pandas as pd
import numpy as np

def home_response(request):
    if request.method == "GET":

        # if request.user=='vyomatesting':
        #     from plotly.offline import plot
        #     from plotly.graph_objs import Scatter

        #     data_pipe = PipeDataProcessed.objects.all()[:200]
        #     x_data = data_pipe.values_list('timestamp', flat=True)
        #     x_data = [x-x_data[0] for x in x_data]
        #     y_data = data_pipe.values_list('weight', flat=True)
        #     y_data = [y-0 for y in y_data ]
        #     plot_div = plot([Scatter(x=x_data, y=y_data, mode='lines', name='test', opacity=0.8, marker_color='green')], output_type='div')
        # else:
        #     plot_div = 0
        #     x_data=0
        #     y_data=0

        machine = Machine.objects.filter(user=request.user)
        machine_id_from_url = request.GET.get('id', None)
        machine_list = [] 
        for i in machine:
            machine_list.append(i.toDic()["machine_id"])
            machine_list.append(i.toDic()["plant_name"]) #append plant name for that machine
        machine_in_plant = Machine.objects.filter(plant_name=machine_list[1])
        m_plant={} #context will be something like m_plant = {"machine_1": 1, "machine_17": 17}
        for m in machine_in_plant:
            v="machine_"+m.machine_id
            m_plant[v] = int(m.machine_id)
        if machine_id_from_url == None:
            machineIdforCalculation = machine_list[0] #when user login assigned machine_id with that login
            curr_machine = int(machine_list[0]) #Selected Machine
        else:
            machineIdforCalculation = machine_id_from_url #when user select machine from frontend
            curr_machine = int(machineIdforCalculation)
        #----------------------------------
        #To plot graph using chart.js
        g_df = PipeDataProcessed.objects.filter(machine_id = machineIdforCalculation)[:500]
        df = pd.DataFrame.from_records(g_df.values())
        df['Pipe Type'] = df['basic_metarial'].astype(str)  + " " + df['standard_type_classification'].astype(str) + " " +  df['pressure_type_specification'].astype(str)
        unique_pipe_type=df['Pipe Type'].unique()
        g_dic={}
        for pipe in unique_pipe_type:
            g_df_pipe_wise = df.loc[df['Pipe Type']==pipe]
            g_df_pipe_wise['site_local_time']=pd.to_datetime(g_df_pipe_wise['site_local_time'], format='%d/%m/%y %H:%M:%S').dt.strftime('%Y-%m-%d %H:%M:%S')
            g_df_pipe_wise['site_local_time'] = g_df_pipe_wise['site_local_time'].astype(str)
            g_pipe_time=g_df_pipe_wise['site_local_time'].tolist()
            g_pipe_weight = g_df_pipe_wise['weight'].tolist()
            g_dic[pipe]=[]
            g_dic[pipe].append(g_pipe_time)
            g_dic[pipe].append(g_pipe_weight)
        g_json = dumps(g_dic)


       
        machinestatus = MachineStatus.objects.filter(machine_id = machineIdforCalculation)
        pipdataprocess = PipeDataProcessed.objects.filter(machine_id= machineIdforCalculation)[:100]
        try:
            data = pipdataprocess[0]
            t0 = data.site_local_time
        except:
            time_stamp=time.time()
            machine_timezone = MachineTimeZone.objects.get(machine_id=machineIdforCalculation)
            no_data_timezone = datetime.datetime.fromtimestamp(int(time_stamp), tz=pytz.timezone(machine_timezone.machine_timezone))
            data = no_data_timezone = no_data_timezone.replace(tzinfo=None)
            t0 = data
        t1 = t0.replace(tzinfo = None)
        t2 = time.time()

        timezone_time = dt.now()
        machine_timezone = MachineTimeZone.objects.get(machine_id = machineIdforCalculation)
        mt = machine_timezone.machine_timezone
        t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
        t4 = t3.replace(tzinfo = None)
        time_difference = (t4 - t1).total_seconds()

        #machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
        #local_time = timezone_time.astimezone(timezone(machine_timezone.machine_timezone))
        enddate = startdate = timezone_time.strftime('%Y-%m-%d')  
        #starttime = (localtime - timedelta(hours=2)).strftime('%H:%M')
        endtime = timezone_time.strftime('%H:%M')
        return 'home.html', {"machinewithstatus" : zip_longest(machine, machinestatus), "pipedata": pipdataprocess,"startdate": enddate, "enddate": enddate, "starttime": endtime, "endtime": endtime,'time_difference' : time_difference, 'g_json': g_json, 'unique_pipe_type': unique_pipe_type, "machine_id": machineIdforCalculation, "m_plant": m_plant, "curr_machine": curr_machine}

    elif request.method == "POST":

        startdate = request.POST['startdate']
        enddate = request.POST['enddate']
        starttime = request.POST['starttime']
        endtime = request.POST['endtime']
        machine = Machine.objects.filter(user=request.user)
        machine_list = []
        for i in machine:
            machine_list.append(i.toDic()["machine_id"])
        machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
        machine_server = MachineLocalServer.objects.get(machine_id=machine_list[0])
        starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")
        endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")
        pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)
        data = PipeDataProcessed.objects.filter(machine_id__in=machine_list)[:100]
        data = data[0]
        t0 = data.site_local_time
        t1 = t0.replace(tzinfo = None)
        t2 = time.time()

        timezone_time = dt.now()
        machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
        mt = machine_timezone.machine_timezone
        t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
        t4 = t3.replace(tzinfo = None)
        time_difference = (t4 - t1).total_seconds()

        return 'home.html', {"machinewithstatus" : zip_longest(machine, machinestatus), "pipedata" : pipdataprocess, "startdate" : startdate, "enddate" : enddate, "starttime": starttime, "endtime": endtime, 'time_difference' : time_difference}



def excel_response(request):
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    starttime = request.GET.get('starttime')
    endtime = request.GET.get('endtime')
    if(startdate == None and enddate == None and starttime and endtime == None):
        return HttpResponse(status = 400)
    machine_list = [] 
    for i in Machine.objects.filter(user=request.user):
        machine_list.append(i.toDic()["machine_id"])

    #machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
    #machine_server = MachineLocalServer.objects.get(machine_id=machine_list[0])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="raw_data_' + startdate + '_' + starttime + '_to_' + enddate + '_' + endtime + '.xlsx"'


    workbook = openpyxl.Workbook()
    sheet = workbook.active

    #columns = ['Pipe Type', 'Outer Diameter', 'Length','Count', 'Weight (Gram)', 'Max Weight (Gram)', 'Min Weight (Gram)', 'Weight Gain (Gram)', 'Status', 'Weighing Time']
    #29_12_2020
    columns = ['Weighing Time','Running Count', 'Pipe Type', 'Outer Diameter', 'Length', ' St Max Weight (Gram)','St Min Weight (Gram)', 'Actual Weight (Gram)','RM Gain/Loss (Gram)','Status']
    sheet_columns = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1']
    for i in range(len(columns)):
        sheet[sheet_columns[i]] = columns[i]
    for col_range in range(1, 11):
        cell_title = sheet.cell(1, col_range)
        cell_title.fill = openpyxl.styles.PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        cell_title.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')

    font = openpyxl.styles.Font(name='Times New Roman',bold=True)
    for i in range(1,11):
        sheet.cell(1,i).font = font
        sheet.cell(1, i).border = Border(left=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    sheet.cell(1, i).border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
    sheet.column_dimensions['A'].width = 24#30
    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['B'].width = 7#16
    sheet.column_dimensions['C'].width = 15#8
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 35#20
    sheet.column_dimensions['G'].width = 16#20
    sheet.column_dimensions['H'].width = 8#20
    sheet.column_dimensions['I'].width = 18#14
    sheet.column_dimensions['J'].width = 20#36

    starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
    starttimestamp = starttimestamp
    endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
    endtimestamp = endtimestamp
    pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)

    row_num = 1
    for row in pipdataprocess:
        row_num += 1
        rowToDict = row.toDic()
        #Pipe Type
        sheet.cell(row_num, 3).value = str(rowToDict["basic_metarial"]) + ' ' + str(rowToDict["standard_type_classification"]) + ' ' + str(rowToDict["pressure_type_specification"])
        sheet.cell(row_num, 3).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 3).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Outer diameter
        sheet.cell(row_num, 4).value = str(rowToDict["outer_diameter"]) + ' ' + str(rowToDict["outer_diameter_unit"])
        sheet.cell(row_num, 4).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 4).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Length
        sheet.cell(row_num, 5).value = str(rowToDict["length"]) + ' ' + str(rowToDict["length_unit"])
        sheet.cell(row_num, 5).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 5).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #count
        sheet.cell(row_num, 2).value = int(rowToDict["count"])
        sheet.cell(row_num, 2).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 2).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weight
        sheet.cell(row_num, 8).value = int(rowToDict["weight"])
        sheet.cell(row_num, 8).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 8).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #max Weight
        sheet.cell(row_num, 6).value = int(rowToDict["maxweight"])
        sheet.cell(row_num, 6).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 6).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #min weight
        sheet.cell(row_num, 7).value = int(rowToDict["minweight"])
        sheet.cell(row_num, 7).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 7).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weight gain
        if rowToDict["weightgain"] == 0:
            sheet.cell(row_num, 9).value = int(rowToDict["weightloss"])
            sheet.cell(row_num, 9).alignment = openpyxl.styles.Alignment(horizontal='center')
            sheet.cell(row_num, 9).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        else:
            sheet.cell(row_num, 9).value = int(rowToDict["weightgain"])
            sheet.cell(row_num, 9).alignment = openpyxl.styles.Alignment(horizontal='center')
            sheet.cell(row_num, 9).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Pass Status
        sheet.cell(row_num, 10).value = str(rowToDict["pass_status"])
        sheet.cell(row_num, 10).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 10).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weighing time
        #Souren//30/03/2021--------changed timezone
        sheet.cell(row_num, 1).value =  timezone.localtime(rowToDict["site_local_time"]).strftime("%b.%d,%Y %I:%M:%S %p")#("%A, %B %d, %Y %I:%M %p")
        sheet.cell(row_num, 1).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 1).border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))

    workbook.save(response)

    return response



