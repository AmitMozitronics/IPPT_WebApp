from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_protect

import datetime, pytz
from datetime import timedelta
from django.utils import timezone

from django.db.models import Sum, Avg, Count
from django.db.models.functions import TruncDay 

import xlwt

import openpyxl
from openpyxl.styles.borders import Border,Side
from django.contrib.gis.geoip2 import GeoIP2
from api.models import *
from . check_time_zone import show_user_time
import pytz
from datetime import datetime as dt
from itertools import zip_longest
import time

def showUserTime(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    g = GeoIP2()
    ip_details = g.city(ip)
    user_time_zone = ip_details['time_zone']
    return user_time_zone

def home_response(request):
    """userTimeZone = showUserTime(request)
    startdate = request.GET.get('startdate', None)
    enddate = request.GET.get('enddate', None)
    starttime = request.GET.get('starttime', None)
    endtime = request.GET.get('endtime', None)
    machine = Machine.objects.filter(user=request.user)
    try:
        last_time = LastSearchTime.objects.get(username = request.user.username)
    except:
        #need to work here later
        pass
    machine_list = [] 
    for i in machine:
        machine_list.append(i.toDic()["machine_id"]) 
    if(startdate == None and enddate == None and starttime == None and endtime == None):
        timezone_localtime = timezone.localtime(timezone.now())
        enddate = startdate = dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d')#timezone_localtime.strftime('%Y-%m-%d')  
        starttime = dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M')#(timezone_localtime - timedelta(hours=2)).strftime('%H:%M')
        endtime = dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M')#timezone_localtime.strftime('%H:%M')
    try:
        starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
        endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
        pipdataprocess = PipeDataProcessed.objects.filter(site_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)
    except Exception as excep:
        timezone_localtime = timezone.localtime(timezone.now())
        enddate = startdate = dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d')#timezone_localtime.strftime('%Y-%m-%d')  
        starttime = dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M')#(timezone_localtime - timedelta(hours=2)).strftime('%H:%M')
        endtime = endtime.strftime('%H:%M')#timezone_localtime.strftime('%H:%M')
        starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
        endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
        pipdataprocess = PipeDataProcessed.objects.filter(site_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list) 
    machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
    from itertools import zip_longest
    return 'home.html', {
        "machinewithstatus": zip_longest(machine, machinestatus), 
        "pipedata": pipdataprocess, 
        "startdate": startdate, "enddate": enddate, 
        "starttime": starttime, "endtime": endtime,
        "issue": str(MetarialIssueShift.objects.filter(shift_date = enddate, user=request.user).count()),
        "duration": str(PipeShiftDuration.objects.filter(shift_date = enddate, user=request.user).count()),
        "notimezone":userTimeZone,
        #"last_time":last_time.start_time.strftime('%H:%M')
    }"""

    if request.method == "GET":
        machine = Machine.objects.filter(user=request.user)
        machine_list = []
        for i in machine:
            machine_list.append(i.toDic()["machine_id"])
        machine_server = MachineLocalServer.objects.get(machine_id=machine_list[0])
        machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
        if machine_server.site_server == "local":
            site_server = machine_server.site_server
            startdate = request.GET.get('startdate')
            enddate = request.GET.get('enddate')
            starttime = request.GET.get('starttime')
            endtime = request.GET.get('endtime')
            if(startdate == None and enddate == None and starttime == None and endtime == None):
                timezone_localtime = timezone.localtime(timezone.now())
                enddate = startdate = timezone_localtime.strftime('%Y-%m-%d')  
                starttime = (timezone_localtime - timedelta(hours=2)).strftime('%H:%M')
                endtime = timezone_localtime.strftime('%H:%M')
            try:
                starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
                endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
                pipdataprocess = PipeDataProcessed.objects.filter(machine_id_in = machine_list)[:10]#(site_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list).order_by('id')[:10]
            except Exception as excep:
                timezone_localtime = timezone.localtime(timezone.now())
                enddate = startdate = timezone_localtime.strftime('%Y-%m-%d')
                starttime = (timezone_localtime - timedelta(hours=2)).strftime('%H:%M')
                endtime = timezone_localtime.strftime('%H:%M')
                starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
                endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")#timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
                pipdataprocess = PipeDataProcessed.objects.filter(machine_id = machine_list[0])[:100]#PipeDataProcessed.objects.filter(site_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list).order_by('-id')[:100]
            machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
            print(starttimestamp, endtimestamp)

            data = pipdataprocess[0]
            t0 = data.site_local_time
            t1 = t0.replace(tzinfo = None)
            t2 = time.time()

            timezone_time = dt.now()
            machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
            mt = machine_timezone.machine_timezone
            t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
            t4 = t3.replace(tzinfo = None)
            time_difference = (t4 - t1).total_seconds()
            

            return 'home.html', {
            "machinewithstatus": zip_longest(machine, machinestatus), 
            "pipedata": pipdataprocess,
            "startdate": startdate, "enddate": enddate, 
            "starttime": starttime, "endtime": endtime,
            "issue": str(MetarialIssueShift.objects.filter(shift_date = enddate, user=request.user).count()),
            "duration": str(PipeShiftDuration.objects.filter(shift_date = enddate, user=request.user).count()),
            "notimezone":show_user_time(request),
            "machine_id":machine_list[0],
            "ss":site_server,
            'time_difference' : time_difference
            #"last_time":last_time.start_time.strftime('%H:%M')
            }


        elif machine_server.site_server == "server":
            site_server = machine_server.site_server
            enddate = startdate = dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d')#timezone_localtime.strftime('%Y-%m-%d')  
            starttime = dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M')#(timezone_localtime - timedelta(hours=2)).strftime('%H:%M')
            endtime = dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M')
            #starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(dt.now()) + ' ' + dt.now() + ':00', "%Y-%m-%d %H:%M:%S"))
            pipdataprocess = PipeDataProcessed.objects.filter(machine_id__in = machine_list)[:100]

            data = pipdataprocess[0]
            t0 = data.site_local_time
            t1 = t0.replace(tzinfo = None)
            t2 = time.time()

            timezone_time = dt.now()
            machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
            mt = machine_timezone.machine_timezone
            t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
            t4 = t3.replace(tzinfo = None)
            time_difference = (t4 - t1).total_seconds()
            return 'home.html', {
            "machinewithstatus": zip_longest(machine, machinestatus), 
            "pipedata": pipdataprocess, 
            "startdate": startdate, "enddate": enddate, 
            "starttime": starttime, "endtime": endtime,
            "issue": str(MetarialIssueShift.objects.filter(shift_date = enddate, user=request.user).count()),
            "duration": str(PipeShiftDuration.objects.filter(shift_date = enddate, user=request.user).count()),
            "notimezone":show_user_time(request),
            "machine_id":machine_list[0],
            "ss":site_server,
            'time_difference' : time_difference
            #"last_time":last_time.start_time.strftime('%H:%M')
            }
    elif request.method == "POST":

        startdate = request.POST['startdate']#dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d'))
        enddate = request.POST['enddate']#GET.get('enddate', dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d'))
        starttime = request.POST['starttime']#GET.get('starttime',dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M') )
        endtime = request.POST['endtime']#GET.get('endtime', dt.now(pytz.timezone(show_user_time(request))).time().strftime('%H:%M'))
        machine = Machine.objects.filter(user=request.user)
        machine_list = []
        for i in machine:
            machine_list.append(i.toDic()["machine_id"])
        machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
        machine_server = MachineLocalServer.objects.get(machine_id=machine_list[0])
        if machine_server.site_server == "local":
        #enddate = startdate = dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d')
        #starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
        #endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
            starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")
            endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")
            pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)
            pipdataprocess1 = PipeDataProcessed.objects.filter(machine_id__in = machine_list)[:10]

            data = pipdataprocess1[0]
            t0 = data.site_local_time
            t1 = t0.replace(tzinfo = None)
            t2 = time.time()

            timezone_time = dt.now()
            machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
            mt = machine_timezone.machine_timezone
            t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
            t4 = t3.replace(tzinfo = None)
            time_difference = (t4 - t1).total_seconds()

            return 'home.html', {
            "machinewithstatus": zip_longest(machine, machinestatus), 
            "pipedata": pipdataprocess, 
            "startdate": startdate, "enddate": enddate, 
            "starttime": starttime, "endtime": endtime,
            "issue": str(MetarialIssueShift.objects.filter(shift_date = enddate, user=request.user).count()),
            "duration": str(PipeShiftDuration.objects.filter(shift_date = enddate, user=request.user).count()),
            "notimezone":show_user_time(request),
            "machine_id":machine_list[0],
            "ss":machine_server.site_server,
            'time_difference' : 'time_difference'
        #"last_time":last_time.start_time.strftime('%H:%M')
        }

        if machine_server.site_server == "server":
            #enddate = startdate = dt.now(pytz.timezone(show_user_time(request))).date().strftime('%Y-%m-%d')
            #starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
            #endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
            starttimestamp = datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S")
            #minutes = datetime.timedelta(hours=5,minutes=30)
            starttimestamp = starttimestamp#-minutes
            endtimestamp = datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S")
            endtimestamp = endtimestamp#-minutes
            pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)
            pipdataprocess1 = PipeDataProcessed.objects.filter(machine_id__in = machine_list)[:10]

            data = pipdataprocess1[0]
            t0 = data.site_local_time
            t1 = t0.replace(tzinfo = None)
            t2 = time.time()

            timezone_time = dt.now()
            machine_timezone = MachineTimeZone.objects.get(machine_id = machine_list[0])
            mt = machine_timezone.machine_timezone
            t3 = datetime.datetime.fromtimestamp(int(t2), tz=pytz.timezone(mt))
            t4 = t3.replace(tzinfo = None)
            time_difference = (t4 - t1).total_seconds()

            return 'home.html', {
            "machinewithstatus": zip_longest(machine, machinestatus), 
            "pipedata": pipdataprocess, 
            "startdate": startdate, "enddate": enddate, 
            "starttime": starttime, "endtime": endtime,
            "issue": str(MetarialIssueShift.objects.filter(shift_date = enddate, user=request.user).count()),
            "duration": str(PipeShiftDuration.objects.filter(shift_date = enddate, user=request.user).count()),
            "notimezone":show_user_time(request),
            "machine_id":machine_list[0],
            "ss":machine_server.site_server,
            'time_difference' : time_difference
        #"last_time":last_time.start_time.strftime('%H:%M')
        }



def excel_response(request):
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    starttime = request.GET.get('starttime')
    endtime = request.GET.get('endtime')
    if(startdate == None and enddate == None and starttime and endtime == None):
        return HttpResponse(status = 400)
    machine_list = [] 
    for i in Machine.objects.filter(user=request.user):
        machine_list.append(i.toDic()["machine_id"])

    #machinestatus = MachineStatus.objects.filter(machine_id__in = machine)
    machine_server = MachineLocalServer.objects.get(machine_id=machine_list[0])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="raw_data_' + startdate + '_' + starttime + '_to_' + enddate + '_' + endtime + '.xlsx"'


    workbook = openpyxl.Workbook()
    sheet = workbook.active

    #columns = ['Pipe Type', 'Outer Diameter', 'Length','Count', 'Weight (Gram)', 'Max Weight (Gram)', 'Min Weight (Gram)', 'Weight Gain (Gram)', 'Status', 'Weighing Time']
    #29_12_2020
    columns = ['Weighing Time','Count','Weight (Gram)','Weight Gain (Gram)','Status','Pipe Type','Outer Diameter','Length','Max Weight (Gram)','Min Weight (Gram)']
    sheet_columns = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1']
    for i in range(len(columns)):
        sheet[sheet_columns[i]] = columns[i]
    for col_range in range(1, 11):
        cell_title = sheet.cell(1, col_range)
        cell_title.fill = openpyxl.styles.PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        cell_title.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')

    font = openpyxl.styles.Font(name='Times New Roman',bold=True)
    for i in range(1,11):
        sheet.cell(1,i).font = font
        sheet.cell(1, i).border = Border(left=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    sheet.cell(1, i).border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
    sheet.column_dimensions['A'].width = 24#30
    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['B'].width = 7#16
    sheet.column_dimensions['C'].width = 15#8
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 35#20
    sheet.column_dimensions['G'].width = 16#20
    sheet.column_dimensions['H'].width = 8#20
    sheet.column_dimensions['I'].width = 18#14
    sheet.column_dimensions['J'].width = 20#36

    if machine_server.site_server == "local":
        starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
        endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
        pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)

    if machine_server.site_server == "server":
        minutes = datetime.timedelta(hours=5,minutes=30)
        starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
        starttimestamp = starttimestamp#-minutes
        endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
        endtimestamp = endtimestamp#-minutes
        pipdataprocess = PipeDataProcessed.objects.filter(site_local_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list)

    row_num = 1
    for row in pipdataprocess:
        row_num += 1
        rowToDict = row.toDic()
        #Pipe Type
        sheet.cell(row_num, 6).value = str(rowToDict["basic_metarial"]) + ' ' + str(rowToDict["standard_type_classification"]) + ' ' + str(rowToDict["pressure_type_specification"])
        sheet.cell(row_num, 6).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 6).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Outer diameter
        sheet.cell(row_num, 7).value = str(rowToDict["outer_diameter"]) + ' ' + str(rowToDict["outer_diameter_unit"])
        sheet.cell(row_num, 7).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 7).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Length
        sheet.cell(row_num, 8).value = str(rowToDict["length"]) + ' ' + str(rowToDict["length_unit"])
        sheet.cell(row_num, 8).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 8).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #count
        sheet.cell(row_num, 2).value = int(rowToDict["count"])
        sheet.cell(row_num, 2).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 2).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weight
        sheet.cell(row_num, 3).value = int(rowToDict["weight"])
        sheet.cell(row_num, 3).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 3).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #max Weight
        sheet.cell(row_num, 9).value = int(rowToDict["maxweight"])
        sheet.cell(row_num, 9).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 9).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #min weight
        sheet.cell(row_num, 10).value = int(rowToDict["minweight"])
        sheet.cell(row_num, 10).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 10).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weight gain
        if rowToDict["weightgain"] == 0:
            sheet.cell(row_num, 4).value = int(rowToDict["weightloss"])
            sheet.cell(row_num, 4).alignment = openpyxl.styles.Alignment(horizontal='center')
            sheet.cell(row_num, 4).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        else:
            sheet.cell(row_num, 4).value = int(rowToDict["weightgain"])
            sheet.cell(row_num, 4).alignment = openpyxl.styles.Alignment(horizontal='center')
            sheet.cell(row_num, 4).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Pass Status
        sheet.cell(row_num, 5).value = str(rowToDict["pass_status"])
        sheet.cell(row_num, 5).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 5).border = Border(left=Side(style='thin'),bottom=Side(style='thin'))
        #Weighing time
        sheet.cell(row_num, 1).value =  timezone.localtime(rowToDict["site_time"]).strftime("%b.%d,%Y %I:%M:%S %p")#("%A, %B %d, %Y %I:%M %p")
        sheet.cell(row_num, 1).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row_num, 1).border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))

    workbook.save(response)
    """wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Raw_Pipe_Data')
    # Sheet header, first row
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['Pipe Type', 'Outer Diameter', 'Length','Count', 'Weight (Gram)', 'Max Weight (Gram)', 'Min Weight (Gram)', 'Weight Gain (Gram)', 'Status', 'Weighing Time']
    row_num = 0
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    starttimestamp = timezone.make_aware(datetime.datetime.strptime(str(startdate) + ' ' + starttime + ':00', "%Y-%m-%d %H:%M:%S"))
    endtimestamp = timezone.make_aware(datetime.datetime.strptime(str(enddate) + ' ' + endtime + ':59', "%Y-%m-%d %H:%M:%S"))
    pipdataprocess = PipeDataProcessed.objects.filter(site_time__range=(starttimestamp, endtimestamp), machine_id__in = machine_list) 
    for row in pipdataprocess:
        row_num += 1
        rowtoDic = row.toDic()
        ws.write(row_num, 0, str(rowtoDic["basic_metarial"]) + ' ' + str(rowtoDic["standard_type_classification"]) + ' ' + str(rowtoDic["pressure_type_specification"]), font_style)
        ws.write(row_num, 1, str(rowtoDic["outer_diameter"]) + ' ' + str(rowtoDic["outer_diameter_unit"]), font_style)
        ws.write(row_num, 2, str(rowtoDic["length"]) + ' ' + str(rowtoDic["length_unit"]), font_style)
        ws.write(row_num, 3, str(rowtoDic["count"]), font_style)
        ws.write(row_num, 4, str(rowtoDic["weight"]), font_style)
        ws.write(row_num, 5, str(rowtoDic["maxweight"]), font_style)
        ws.write(row_num, 6, str(rowtoDic["minweight"]), font_style)
        if rowtoDic["weightgain"] == 0:
            ws.write(row_num, 7, str(rowtoDic["weightloss"]), font_style)
        else:
            ws.write(row_num, 7, str(rowtoDic["weightgain"]), font_style)
        ws.write(row_num, 8, str(rowtoDic["pass_status"]), font_style)
        ws.write(row_num, 9, timezone.localtime(rowtoDic["site_time"]).strftime("%A, %B %d, %Y %I:%M %p"), font_style)
    wb.save(response)"""

    return response

